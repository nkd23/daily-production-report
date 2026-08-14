"""Excel export that reproduces the factory's original daily report layout:

Line | NEW OUT-TAR | NEW EFF-TAR | Buyer | SAM | Shift | OUT-SEW | EFF-SEW |
OUT-FIN(ScanPack) | OUT-FIN(Fin) | EFF-FIN | VAR | Tồn Dip | Tồn trước PI |
Issue/Lí do

(The original sheet had a single "WIP FIN" column here; the factory split it
into the two stock readings the Tổ trưởng actually reports at end of shift.)

Rows: one per Line, grouped by Executive (subtotal row after each group),
then by PU1/PU2 (subtotal row after each PU), then a final TTL row.
"""

from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas import ExecutiveSummary, IssueItem, KpiSummary, LineDaySummary

# NOTE: these thresholds are NOT confirmed with the factory yet — they are a
# best guess reproducing the "tô cam / tô hồng" warning look of the original
# file. Adjust here once IE/PPC confirms the real cutoffs. Comparison is the
# ratio of actual efficiency to that line's own target efficiency.
EFF_WARNING_THRESHOLD = 0.75  # ratio actual/target below this -> tô cam (orange)
EFF_CRITICAL_THRESHOLD = 0.60  # ratio actual/target below this -> tô hồng/tím nhạt (pink)

COLUMNS = [
    ("Line", 12),
    ("NEW OUT-TAR", 13),
    ("NEW EFF-TAR", 13),
    ("Buyer", 14),
    ("SAM", 9),
    ("Shift", 8),
    ("OUT-SEW", 11),
    ("EFF-SEW", 11),
    ("OUT-FIN(ScanPack)", 16),
    ("OUT-FIN(Fin)", 13),
    ("EFF-FIN", 11),
    ("VAR", 10),
    ("Tồn Dip", 11),
    ("Tồn trước PI", 13),
    ("Issue/Lí do", 40),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
EXEC_FILL = PatternFill("solid", fgColor="DCE6F1")
PU_FILL = PatternFill("solid", fgColor="B8CCE4")
TTL_FILL = PatternFill("solid", fgColor="1F4E78")
TTL_FONT = Font(color="FFFFFF", bold=True)
WARNING_FILL = PatternFill("solid", fgColor="FFC000")
CRITICAL_FILL = PatternFill("solid", fgColor="F4B7C4")
VAR_NEGATIVE_FONT = Font(color="C00000", bold=True)
THIN_BORDER = Border(*(Side(style="thin", color="BFBFBF") for _ in range(4)))


def _avg(values: list[float]) -> float | None:
    values = [v for v in values if v is not None]
    return round(sum(values) / len(values), 2) if values else None


def _shift_weighted_avg(pairs: list[tuple[float | None, float | None]]) -> float | None:
    """Weighted average of a percentage, weighted by how many shifts each line
    ran. Matches app.services.aggregation._shift_weighted_avg (keep in sync)
    and the original sheet: SUMPRODUCT(Shift, EFF%) / SUM(Shift)."""
    total_weight = 0.0
    total_weighted = 0.0
    for weight, eff in pairs:
        if not weight or eff is None:
            continue
        total_weight += weight
        total_weighted += weight * eff
    if not total_weight:
        return None
    return round(total_weighted / total_weight, 2)


def _eff_fill(actual: float | None, target: float | None):
    if actual is None or not target:
        return None
    ratio = actual / target
    if ratio < EFF_CRITICAL_THRESHOLD:
        return CRITICAL_FILL
    if ratio < EFF_WARNING_THRESHOLD:
        return WARNING_FILL
    return None


def _write_row(ws, row_idx: int, values: list, *, bold=False, fill=None, font=None, number_formats=None):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = THIN_BORDER
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        elif bold:
            cell.font = Font(bold=True)
        if number_formats and number_formats.get(col_idx):
            cell.number_format = number_formats[col_idx]
    return row_idx


# Keyed by 1-based column index into COLUMNS above.
NUMBER_FORMATS = {
    2: "#,##0",       # NEW OUT-TAR
    3: "0.0\"%\"",    # NEW EFF-TAR
    5: "0.00",        # SAM
    7: "#,##0",       # OUT-SEW
    8: "0.0\"%\"",    # EFF-SEW
    9: "#,##0",       # OUT-FIN(ScanPack)
    10: "#,##0",      # OUT-FIN(Fin)
    11: "0.0\"%\"",   # EFF-FIN
    12: "#,##0",      # VAR
    13: "#,##0",      # Tồn Dip
    14: "#,##0",      # Tồn trước PI
}

EFF_SEW_COL = 8
EFF_FIN_COL = 11
VAR_COL = 12
ISSUE_COL = 15


def _line_row_values(line: LineDaySummary) -> list:
    # 0 means the line hasn't been given a real NEW OUT-TAR/NEW EFF-TAR yet
    # (management supplies these weekly, on Mondays) - leave the cell blank
    # rather than showing a misleading "0".
    return [
        line.line_number,
        line.target_output or None,
        float(line.target_eff) if line.target_eff else None,
        line.buyer,
        float(line.sam),
        line.shift_display,
        line.out_sew,
        float(line.eff_sew) if line.eff_sew is not None else None,
        line.out_fin_scanpack,
        line.out_fin_fin,
        float(line.eff_fin) if line.eff_fin is not None else None,
        line.var,
        line.wip_dip,
        line.wip_pre_pi,
        line.issue_note or "",
    ]


def _subtotal_row_values(label: str, items: list[LineDaySummary], sam_avg: float | None) -> list:
    """Mirrors app.services.aggregation._build_group_summary - keep in sync.
    SAM is passed in because the original sheet cascades it from the level
    below rather than recomputing it from the raw lines."""
    target_output_sum = sum(i.target_output for i in items)
    return [
        label,
        target_output_sum or None,
        _shift_weighted_avg([(i.shift_weight, float(i.target_eff)) for i in items if i.target_eff]),
        "",
        sam_avg,
        sum(i.shift_weight for i in items) or "",
        sum((i.out_sew or 0) for i in items),
        _shift_weighted_avg([(i.eff_sew_weight, float(i.eff_sew) if i.eff_sew is not None else None) for i in items]),
        sum((i.out_fin_scanpack or 0) for i in items),
        sum((i.out_fin_fin or 0) for i in items),
        _shift_weighted_avg([(i.eff_fin_weight, float(i.eff_fin) if i.eff_fin is not None else None) for i in items]),
        sum((i.var or 0) for i in items),
        sum((i.wip_dip or 0) for i in items),
        sum((i.wip_pre_pi or 0) for i in items),
        "",
    ]


KPI_LABEL_FONT = Font(bold=True, size=9, color="64748B")
KPI_VALUE_FONT = Font(bold=True, size=16, color="1F4E78")


def _add_dashboard_sheet(
    wb: Workbook,
    report_date: date,
    kpi: KpiSummary,
    executives: list[ExecutiveSummary],
    issues: list[IssueItem],
):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value=f"DASHBOARD SẢN LƯỢNG - {report_date.strftime('%d/%m/%Y')}").font = Font(
        bold=True, size=16, color="1F4E78"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    kpi_items = [
        ("Target Output", f"{kpi.total_target_output:,}"),
        ("Actual Output", f"{kpi.total_actual_output:,}"),
        ("% Hoàn thành", f"{kpi.completion_rate}%" if kpi.completion_rate is not None else "-"),
        ("EFF SEW TB", f"{kpi.avg_eff_sew}%" if kpi.avg_eff_sew is not None else "-"),
        ("EFF FIN TB", f"{kpi.avg_eff_fin}%" if kpi.avg_eff_fin is not None else "-"),
        ("Line có Issue", f"{kpi.lines_with_issue} ({kpi.lines_submitted}/{kpi.lines_total} đã nộp)"),
    ]
    for col_idx, (label, value) in enumerate(kpi_items, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.cell(row=3, column=col_idx, value=label).font = KPI_LABEL_FONT
        ws.cell(row=4, column=col_idx, value=value).font = KPI_VALUE_FONT

    # Small data table (Executive: Target vs Thực tế) that both charts read from.
    data_row = 7
    ws.cell(row=data_row, column=1, value="Executive").font = Font(bold=True)
    ws.cell(row=data_row, column=2, value="Target").font = Font(bold=True)
    ws.cell(row=data_row, column=3, value="Thực tế").font = Font(bold=True)
    for offset, exec_summary in enumerate(executives, start=1):
        r = data_row + offset
        ws.cell(row=r, column=1, value=exec_summary.executive_name)
        ws.cell(row=r, column=2, value=exec_summary.target_output)
        ws.cell(row=r, column=3, value=exec_summary.out_fin_fin)
    data_end_row = data_row + len(executives)

    if executives:
        cats = Reference(ws, min_col=1, min_row=data_row + 1, max_row=data_end_row)

        bar = BarChart()
        bar.type = "col"
        bar.title = "Output theo Executive (Target vs Thực tế)"
        bar.y_axis.title = "Sản lượng"
        bar_data = Reference(ws, min_col=2, max_col=3, min_row=data_row, max_row=data_end_row)
        bar.add_data(bar_data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width = 17
        bar.height = 9
        ws.add_chart(bar, f"A{data_end_row + 2}")

        pie = PieChart()
        pie.title = "Tỷ trọng Output thực tế theo Executive"
        pie_data = Reference(ws, min_col=3, min_row=data_row, max_row=data_end_row)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(cats)
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.width = 13
        pie.height = 9
        ws.add_chart(pie, f"H{data_end_row + 2}")

    issue_row = data_end_row + 21
    ws.cell(row=issue_row, column=1, value="Issue trong ngày").font = Font(bold=True, size=12)
    issue_row += 1
    for col_idx, header in enumerate(["Line", "Buyer", "Executive", "Issue / Lí do"], start=1):
        cell = ws.cell(row=issue_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.column_dimensions["D"].width = 60
    for issue in issues:
        issue_row += 1
        ws.cell(row=issue_row, column=1, value=issue.line_number)
        ws.cell(row=issue_row, column=2, value=issue.buyer)
        ws.cell(row=issue_row, column=3, value=issue.executive_name)
        ws.cell(row=issue_row, column=4, value=issue.issue_note).alignment = Alignment(wrap_text=True)
    if not issues:
        issue_row += 1
        ws.cell(row=issue_row, column=1, value="Không có issue nào được ghi nhận.")


def generate_daily_excel(
    lines: list[LineDaySummary],
    report_date: date,
    kpi: KpiSummary | None = None,
    executives: list[ExecutiveSummary] | None = None,
    issues: list[IssueItem] | None = None,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = report_date.isoformat()

    ws.cell(row=1, column=1, value=f"BÁO CÁO SẢN LƯỢNG NGÀY {report_date.strftime('%d/%m/%Y')}").font = Font(
        bold=True, size=14
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    header_row = 3
    for col_idx, (title, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _write_row(ws, header_row, [c[0] for c in COLUMNS], fill=HEADER_FILL, font=HEADER_FONT)
    ws.freeze_panes = f"A{header_row + 1}"

    row_idx = header_row + 1

    lines_by_pu: dict[str, list[LineDaySummary]] = {}
    for line in lines:
        lines_by_pu.setdefault(line.pu_group.value, []).append(line)

    pu_sams: list[float | None] = []
    for pu_group in sorted(lines_by_pu.keys()):
        pu_lines = lines_by_pu[pu_group]
        lines_by_exec: dict[str, list[LineDaySummary]] = {}
        for line in pu_lines:
            lines_by_exec.setdefault(line.executive_name, []).append(line)

        exec_sams: list[float | None] = []
        for executive_name, exec_lines in lines_by_exec.items():
            for line in exec_lines:
                _write_row(ws, row_idx, _line_row_values(line), number_formats=NUMBER_FORMATS)
                eff_sew_fill = _eff_fill(float(line.eff_sew) if line.eff_sew is not None else None, float(line.target_eff))
                if eff_sew_fill:
                    ws.cell(row=row_idx, column=EFF_SEW_COL).fill = eff_sew_fill
                eff_fin_fill = _eff_fill(float(line.eff_fin) if line.eff_fin is not None else None, float(line.target_eff))
                if eff_fin_fill:
                    ws.cell(row=row_idx, column=EFF_FIN_COL).fill = eff_fin_fill
                if line.var is not None and line.var < 0:
                    ws.cell(row=row_idx, column=VAR_COL).font = VAR_NEGATIVE_FONT
                row_idx += 1

            exec_sam = _avg([float(l.sam) for l in exec_lines])
            exec_sams.append(exec_sam)
            _write_row(
                ws,
                row_idx,
                _subtotal_row_values(f"{executive_name} - TTL", exec_lines, exec_sam),
                fill=EXEC_FILL,
                bold=True,
                number_formats=NUMBER_FORMATS,
            )
            row_idx += 1

        pu_sam = _avg(exec_sams)
        pu_sams.append(pu_sam)
        _write_row(
            ws,
            row_idx,
            _subtotal_row_values(f"{pu_group} - TTL", pu_lines, pu_sam),
            fill=PU_FILL,
            bold=True,
            number_formats=NUMBER_FORMATS,
        )
        row_idx += 1

    _write_row(
        ws,
        row_idx,
        _subtotal_row_values("TTL", lines, _avg(pu_sams)),
        fill=TTL_FILL,
        font=TTL_FONT,
        number_formats=NUMBER_FORMATS,
    )

    ws.cell(row=header_row, column=1).alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=header_row, max_row=row_idx, min_col=ISSUE_COL, max_col=ISSUE_COL):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    if kpi is not None and executives is not None and issues is not None:
        _add_dashboard_sheet(wb, report_date, kpi, executives, issues)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
